import openpyxl
import os
from openpyxl.styles import PatternFill

base_path = os.path.abspath(os.path.dirname(os.getcwd()))


class HandExcel():
    """获取，处理excel中各种数据"""

    def load_excel(self):
        """加载excel对象"""
        excel_path = os.path.join(base_path, "Case", "shopV.xlsx")
        open_excel = openpyxl.load_workbook(excel_path)
        return open_excel

    def get_sheet_data(self, index=None):
        """加载所有sheet内容,默认为第一个sheet"""

        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self, row, cols):
        """获取某一单元格数据"""
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        """获取行数"""
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """获取某一行的数据"""
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value, color=None):
        """数据写入excel某个单元格内"""
        wb = self.load_excel()
        wr = wb.active

        wr.cell(row, cols, value)
        if color == "red":
            "失败case 的结果加粗并标红"

            red_fill = PatternFill("solid", fgColor="FF0000")
            d6 = wr['M' + str(row)]
            d6.fill = red_fill

        excel_path = os.path.join(base_path, "Case", "shopV.xlsx")
        wb.save(excel_path)

    def get_columns_value(self, key=None):
        """获取某一列的数据，默认第一列"""
        columns_list = []
        if key is None:
            key = "A"
        for i in self.get_sheet_data()[key]:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        """根据case名称获取行号"""
        num = 1
        cols_data = self.get_columns_value()
        for clo_data in cols_data:
            if case_id == clo_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        """获取exccel里面所有的数据(列表嵌套展示)"""
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))

        return data_list[:-1]


excel_data = HandExcel()

if __name__ == '__main__':
    excel_data = HandExcel()

    print(excel_data.get_excel_data())
